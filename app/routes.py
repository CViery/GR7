from app import app
from flask import request, redirect, render_template, flash, session, jsonify, make_response, Response
from services import login, cadastrar_notas, cadastrar_duplicata, dados_notas, faturamento, utills, xlxs, rotas
from datetime import datetime
from flask_paginate import Pagination, get_page_parameter
from database import gastos_db, conection
from xhtml2pdf import pisa
from io import BytesIO
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image



PERMISSAO_TOTAL_ADMIN = 0
PERMISSAO_GR7_USER = 1
PERMISSAO_PORTAL_ADMIN = 2
SEM_PERMISSAO = 3
PERMISSAO_GR7_MORUMBI_ADMIN = 4



class Routes:
    def __init__(self):
        pass

    @app.route('/')
    def show_login():
        db = login.Login()
        empresas = db.empresas()

        return render_template(
            'login.html',
            empresas=empresas
        )


    @app.route('/autenticar', methods=['POST'])
    def autenticar():
        try:
            usuario = request.form.get('usuario', '').strip()
            senha = request.form.get('senha', '')
            empresa_form = request.form.get('empresa')

            if not usuario or not senha or not empresa_form:
                flash('Preencha todos os campos!')
                return redirect('/')

            try:
                empresa = int(empresa_form)

            except (TypeError, ValueError):
                flash('Empresa inválida.')
                return redirect('/')

            login_service = login.Login()

            resultado = login_service.login(
                usuario.upper(),
                senha
            )

            if not resultado['autenticado']:
                flash(resultado['mensagem'])
                return redirect('/')

            guia_usuario = resultado['guia']

            # Guia 0 possui acesso a todas as empresas.
            tem_acesso = (
                guia_usuario == 0
                or guia_usuario == empresa
            )

            if not tem_acesso:
                session.clear()

                flash(
                    'Desculpe, seu acesso não permite '
                    'acessar esta empresa.'
                )
                return redirect('/')

            session['empresa'] = empresa

            return redirect('/home')

        except Exception as erro:
            print(f'Erro durante autenticação: {erro}')

            session.clear()

            flash('Ocorreu um erro. Tente novamente.')
            return redirect('/')

    @app.route('/logout')
    def logout():
        # Remover todas as variáveis de sessão relacionadas ao usuário
        session.pop('usuario', None)
        session.pop('empresa', None)
        session.pop('permission_empresa', None)  # Caso esteja utilizando esta variável
        flash('Você saiu da sessão com sucesso.')
        return redirect('/')

    
    @app.route('/home')
    def home():
        try:
            # Verificar se o usuário está logado
            if 'usuario' not in session:
                flash('Usuário não está logado.')
                return redirect('/')

            # Obter as informações da sessão
            usuario = session['usuario']
            empresa = session['empresa']
            permission = session.get('permission', None)

            if empresa == 1:
                if permission == 'ADMIN':
                    return rotas.render_gr7_admin(usuario)
                else:
                    flash('Você não tem permissão para acessar esta página.')
                    return redirect('/')

            if empresa == 2:
                if permission == 'ADMIN':
                    return rotas.render_portal_admin(usuario)
                elif permission == 'NORMAL':
                    return rotas.render_portal_normal(usuario)
                else:
                    flash('Permissão de acesso inválida.')
                    return redirect('/')
                
            if empresa == 3:
                if permission == 'ADMIN':
                    return rotas.render_gr7_morumbi_admin(usuario)
                elif permission == 'NORMAL':
                    return rotas.render_gr7_morumbi_normal(usuario)
                else:
                    flash('Permissão de acesso inválida.')
                    return redirect('/')

            else:
                flash('Empresa não reconhecida.')
                return redirect('/')

        except Exception as e:
            print(f"Erro no carregamento da página home: {e}")
            flash('Ocorreu um erro ao carregar a página. Tente novamente.')
            return redirect('/')
        

    @app.route('/faturamento', methods=['GET', 'POST'])
    def tela_faturamentos():
        if 'usuario' not in session:
            print('Usuário não está logado')
            return redirect('/')

        empresa = session.get('empresa')
        permissao = session.get('permission')
        db = faturamento.Faturamento()
        services = utills.Utills()
        
        def get_mes_nome(codigo):
            meses_map = {
                '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março', '04': 'Abril',
                '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
                '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro',
                '1': 'Janeiro', '2': 'Fevereiro', '3': 'Março', '4': 'Abril',
                '5': 'Maio', '6': 'Junho', '7': 'Julho', '8': 'Agosto',
                '9': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
            }
            return meses_map.get(str(int(codigo)).zfill(2), "Mês Inválido")

        
        meses = [(str(i).zfill(2), get_mes_nome(str(i))) for i in range(1, 13)]
        anos = [str(ano) for ano in range(2024, 2031)]

        
        if request.method == 'POST':
            mes_dados = request.form.get('mes')
            ano_dados = request.form.get('ano')
        else:
            now = datetime.now()
            mes_dados = now.strftime('%m')
            ano_dados = now.strftime('%Y')

       
        if mes_dados and ano_dados:
            session['mes_atual'] = mes_dados
            session['ano_atual'] = ano_dados

        try:
            
            valor_faturamento_total = db.faturamento_total_mes(mes_dados, ano_dados)
            valor_faturamento_meta = db.faturamento_meta_mes(mes_dados, ano_dados)
            faturamento_mecanicos = db.faturamento_mecanico(mes_dados, ano_dados)
            faturamento_cias = db.faturamento_companhia(mes_dados, ano_dados)
            faturamento_servico = db.faturamento_servico(mes_dados, ano_dados)
            valor_dinheiro = db.faturamento_dinheiro(mes_dados, ano_dados)
            ticket = services.ticket(mes_dados, ano_dados)
            passagens = services.passagens(mes_dados, ano_dados)
            valor_meta_int = db.faturamento_meta_mes_int(mes_dados, ano_dados)
            mes_select = get_mes_nome(mes_dados)
            ano_select = ano_dados

            return render_template(
                'faturamentos.html',
                anos=anos,
                meses=meses,
                valor_faturamento_total=valor_faturamento_total,
                valor_faturamento_meta=valor_faturamento_meta,
                faturamento_mecanicos=faturamento_mecanicos,
                faturamento_companhia=faturamento_cias,
                faturamento_servico=faturamento_servico,
                empresa=empresa,
                valor_dinheiro=valor_dinheiro,
                ticket=ticket,
                passagens=passagens,
                valor_meta_int=valor_meta_int,
                mes_escolhido=mes_select,
                ano_escolhido=ano_select
            )

        except Exception as e:
            print(f"Erro ao processar faturamento: {e}")
            return "Erro interno no servidor", 500



    @app.route('/faturamentos/cadastrar', methods=['GET'])
    def cadastrar_faturamento():
        if 'usuario' not in session:
            return redirect('/')

        db = faturamento.Faturamento()
        empresa = session['empresa']
        if db is None:
            return redirect('/')

        return render_template(
            'cadastrar_faturamento.html',
            empresa=session['empresa'],
            cias=db.companhias(),
            mecanicos=db.funcionarios(empresa),
            response=''
        )


    @app.route('/submit_form', methods=['POST'])
    def submit_form():
        if 'usuario' not in session:
            return redirect('/')

        db = faturamento.Faturamento()

        if db is None:
            return redirect('/')

        data = request.form.to_dict()
        usuario = session['usuario']

        ja_existe = db.cadastrar(data, usuario)
        empresa = session['empresa']
        if ja_existe:
            response = f"A OS {data['num_os']} já está cadastrada"
        else:
            response = f"A OS {data['num_os']} cadastrada com sucesso"

        return render_template(
            'cadastrar_faturamento.html',
            empresa=session['empresa'],
            cias=db.companhias(),
            mecanicos=db.funcionarios(empresa),
            response=response
        )       

    @app.route('/faturamentos/consultar', methods=['GET', 'POST'])
    def consultar_faturamentos():
        if 'usuario' in session:
            empresa = session['empresa']
            db = faturamento.Faturamento()
            cias = db.companhias()
            mecanicos = db.funcionarios(empresa)
            if request.method == 'POST':
                data_inicio = request.form.get('data_inicio')
                data_fim = request.form.get('data_fim')
                companhia = request.form.get('companhia')
                numero_os = request.form.get('num_os')
                    
                placa = request.form.get('placa')
                mecanico_servico = request.form.get('mecanico_servico')

                # Implementar a lógica para buscar os faturamentos no banco de dados com base nos filtros
                faturamentos = db.filtrar_os(
                        data_inicio, data_fim, placa, mecanico_servico, numero_os, companhia)
                    
                valor = db.filtrar_os_valor(data_inicio, data_fim, placa, mecanico_servico, numero_os, companhia)
                valor_meta = db.filtrar_os_valor_meta(data_inicio, data_fim, placa, mecanico_servico, numero_os, companhia)
                        
            else:
                    # Se for uma requisição GET, buscar todos os faturamentos ou usar uma lógica padrão
                faturamentos = db.faturamentos_gerais()
                valor = db.faturamentos_gerais_valor()
                valor_meta = db.faturamentos_gerais_valor_meta()
                    
                        

            if faturamentos is None:
                faturamentos = []

            return render_template('consultar_faturamento.html',
                                       empresa=empresa,
                                       cias=cias,
                                       mecanicos=mecanicos,
                                       faturamentos=faturamentos, valor=valor, valor_meta=valor_meta)
        else:
            return redirect('/')
            
    @app.route('/faturamentos/ordens_com_dinheiro/', methods=['GET', 'POST'])
    def consultar_faturamentos_c_dinheiro():
        if 'usuario' in session:
            db = faturamento.Faturamento()
            mes = session.get('mes_atual')
            ano = session.get('ano_atual')
            empresa = session['empresa']
                # Certifique-se de passar a conexão com o banco de dados
                
                
                # Listas para preencher os selects
                

            faturamentos = db.faturamento_dinheiro_ordens(mes,ano)
            
            valor = db.faturamento_dinheiro(mes,ano)
            if faturamentos is None:
                faturamentos = []

            return render_template('consultar_faturamento_dinheiro.html',
                                       empresa=empresa,
                                       faturamentos=faturamentos, valor=valor)
        else:
            return redirect('/')

    @app.route('/baixar')
    def baixar_excel():
        if 'usuario' in session:
            if 'dados_exportar' in session:
                dados = session['dados_exportar']
                excel = xlxs.GerarExcel()
                arquivo = excel.exportar_faturamentos_excel(dados)
                
                if arquivo:
                    return arquivo
                else:
                    print('Erro ao gerar o arquivo Excel.')
                    return redirect('/faturamentos/consultar')
            else:
                print('Dados para exportação não encontrados na sessão.')
                return redirect('/faturamentos/consultar')
        else:
            print('Usuário não está logado.')
            return redirect('/')

    @app.route('/relatorios')
    def page_relatorios():
        
        db = faturamento.Faturamento()
        db_utils = utills.Utills()
        
        mecanicos = db.funcionarios(session['empresa'])
        context = {
                'mecanicos': mecanicos
            } 
        return render_template('relatorios.html', **context)

    @app.route('/fechamento_mensal', methods=['GET', 'POST'])
    def gerar_pdf():
        db = faturamento.Faturamento()
        services = utills.Utills()
        empresa = session['empresa']
        mes = request.form.get('mes')
        ano = request.form.get('ano')
        # Dados para o template
        
        
        valor_faturamento_total = db.faturamento_total_mes(mes, ano)
        valor_faturamento_meta = db.faturamento_meta_mes(mes, ano)
        faturamento_mecanicos = db.faturamento_mecanico(mes, ano)
        faturamento_cias = db.faturamento_companhia(mes, ano)
        faturamento_servico = db.faturamento_servico(mes, ano)
        valor_dinheiro = db.faturamento_dinheiro(mes, ano)
        ticket = services.ticket(mes, ano)
        passagens = services.passagens(mes, ano)
        valor_meta_int = db.faturamento_meta_mes_int(mes, ano)
        mecanicos = db.faturamento_mecanico(mes, ano)
        dados_filtros = db.filtros_mecanico(mes, ano)
        dados_revitalizacao = db.revitalizacao_mecanico(mes,ano)
        context = {
            'valor_faturamento_total': valor_faturamento_total,
            'valor_faturamento_meta': valor_faturamento_meta,
            'faturamento_mecanicos': faturamento_mecanicos,
            'faturamento_companhia': faturamento_cias,
            'faturamento_servico': faturamento_servico,
            'empresa': empresa,
            'valor_dinheiro': valor_dinheiro,
            'ticket': ticket,
            'passagens': passagens,
            'valor_meta_int': valor_meta_int,
            'mes_escolhido': mes,
            'ano_escolhido': ano,
            'dados_filtros': dados_filtros,
            'dados_revitalizacao': dados_revitalizacao
        }
        
        # Renderiza o template HTML
        html = render_template('relatorio_faturamento.html', **context)
        
        # Gera um buffer em memória para o PDF
        pdf_buffer = BytesIO()
        
        # Gera o PDF
        pisa_status = pisa.CreatePDF(
            html.encode('utf-8'), dest=pdf_buffer, encoding='utf-8'
        )
        
        # Verifica se ocorreu um erro
        if pisa_status.err:
            return "Erro ao gerar o PDF", 500

        # Nome seguro para o arquivo
        name_arquivo = f"Relatorio_{mes}_{ano}".replace(" ", "_").replace("/", "-")
        
        # Movendo o ponteiro do buffer para o início
        pdf_buffer.seek(0)
        
        # Retorna o PDF como uma resposta Flask
        response = Response(
            pdf_buffer,
            content_type='application/pdf'
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{name_arquivo}.pdf"'
        return response

   

    @app.route('/fechamento_filtros', methods=['GET', 'POST'])
    def gerar_relatorio_filtros():
        # Obtendo os valores do formulário
        db = faturamento.Faturamento()
        services = utills.Utills()
        empresa = session['empresa']

        mes = request.form.get('mes')
        ano = request.form.get('ano')
        mecanico = request.form.get('mecanico')
        
      


        # Buscando os dados de acordo com os filtros
        dados = db.ordens_filtro_e_higienizacao(mes, ano, mecanico)
        
        # Definindo o nome da empresa
        empresa = session['empresa']

        # Contexto que será passado para o template
        context = {
            'empresa': empresa,
            'mecanico': mecanico,
            'mes': mes,
            'ano': ano,
            'dados': dados
        }

        # Renderizando o template HTML
        html = render_template('relatorio_filtros.html', **context)

        # Criando um buffer de memória para o PDF
        pdf_buffer = BytesIO()

        # Gerando o PDF a partir do HTML
        pisa_status = pisa.CreatePDF(
            html.encode('utf-8'), dest=pdf_buffer, encoding='utf-8'
        )

        # Verificando se ocorreu algum erro na geração do PDF
        if pisa_status.err:
            return "Erro ao gerar o PDF", 500

        # Criando um nome seguro para o arquivo
        name_arquivo = f"Relatorio_{mes}_{ano}_{mecanico}".replace(" ", "_").replace("/", "-")

        # Movendo o ponteiro do buffer para o início
        pdf_buffer.seek(0)

        # Retornando o PDF como resposta Flask
        response = Response(
            pdf_buffer,
            content_type='application/pdf'
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{name_arquivo}.pdf"'

        return response

    @app.route('/fechamento_revitalizacao', methods=['GET', 'POST'])
    def gerar_relatorio_revitalizacao():
        # Obtendo os valores do formulário
        db = faturamento.Faturamento()
        empresa = session['empresa']

        mes = request.form.get('mes')
        ano = request.form.get('ano')
        mecanico = request.form.get('mecanico')
        
        # Buscando os dados de acordo com os filtros
        dados = db.ordens_revitalizacao(mes, ano, mecanico)
        
        # Definindo o nome da empresa
        empresa = session['empresa']

        # Contexto que será passado para o template
        context = {
            'empresa': empresa,
            'mecanico': mecanico,
            'mes': mes,
            'ano': ano,
            'dados': dados
        }

        # Renderizando o template HTML
        html = render_template('relatorio_revitalizacao.html', **context)

        # Criando um buffer de memória para o PDF
        pdf_buffer = BytesIO()

        # Gerando o PDF a partir do HTML
        pisa_status = pisa.CreatePDF(
            html.encode('utf-8'), dest=pdf_buffer, encoding='utf-8'
        )

        # Verificando se ocorreu algum erro na geração do PDF
        if pisa_status.err:
            return "Erro ao gerar o PDF", 500

        # Criando um nome seguro para o arquivo
        name_arquivo = f"Relatorio_{mes}_{ano}_{mecanico}".replace(" ", "_").replace("/", "-")

        # Movendo o ponteiro do buffer para o início
        pdf_buffer.seek(0)

        # Retornando o PDF como resposta Flask
        response = Response(
            pdf_buffer,
            content_type='application/pdf'
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{name_arquivo}.pdf"'

        return response
    
    @app.route('/fechamento_dinheiro', methods=['GET', 'POST'])
    def gerar_relatorio_dinheiro():
        # Obtendo os valores do formulário
        mes = request.form.get('mes')
        ano = request.form.get('ano')
        mecanico = request.form.get('mecanico')
        

        # Inicializando os objetos para acessar os dados
        db = faturamento.Faturamento()
        empresa = session['empresa']

        # Buscando os dados de acordo com os filtros
        dados = db.ordens_dinheiro_relat(mes, ano)
        

        # Contexto que será passado para o template
        context = {
            'empresa': empresa,
            'mecanico': mecanico,
            'mes': mes,
            'ano': ano,
            'dados': dados
        }

        # Renderizando o template HTML
        html = render_template('relatorio_dinheiro.html', **context)

        # Criando um buffer de memória para o PDF
        pdf_buffer = BytesIO()

        # Gerando o PDF a partir do HTML
        pisa_status = pisa.CreatePDF(
            html.encode('utf-8'), dest=pdf_buffer, encoding='utf-8'
        )

        # Verificando se ocorreu algum erro na geração do PDF
        if pisa_status.err:
            return "Erro ao gerar o PDF", 500

        # Criando um nome seguro para o arquivo
        name_arquivo = f"Relatorio_{mes}_{ano}_{mecanico}".replace(" ", "_").replace("/", "-")

        # Movendo o ponteiro do buffer para o início
        pdf_buffer.seek(0)

        # Retornando o PDF como resposta Flask
        response = Response(
            pdf_buffer,
            content_type='application/pdf'
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{name_arquivo}.pdf"'

        return response

    @app.route('/baixar_os', methods=['GET', 'POST'])
    def gerar_relatorio_ordens():
        # Obtendo os valores do formulário
        mes = request.form.get('mes')
        ano = request.form.get('ano')

        
        # Inicializando os objetos para acessar os dados
        db = faturamento.Faturamento()
        empresa = session['empresa']
        

        # Buscando os dados de acordo com os filtros
        dados = db.faturamentos_ordens(mes, ano)

        # Criando o arquivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"Relatório ordens - {empresa}"

        # Adicionando cabeçalhos
        colunas = [
            'Placa', 'Modelo Veículo', 'Data Orçamento', 'Data Faturamento', 'Dias Serviço', 
            'Número OS', 'Companhia', 'Valor Peças', 'Valor Serviços', 'Total OS', 
            'Valor Revitalização', 'Valor Aditivo', 'Quantidade Litros', 'Valor Fluido Sangria',
            'Valor Palheta', 'Valor Limpeza Freios', 'Valor Pastilha Para-brisa', 
            'Valor Filtro', 'Valor Pneu', 'Valor Bateria', 'Modelo Bateria', 
            'Litros Óleo Motor', 'Valor Litro Óleo', 'Marca e Tipo Óleo', 'Mecânico Serviço', 
            'Serviço Filtro', 'Valor P Meta', 'Valor em Dinheiro', 'Valor Serviço Freios', 
            'Valor Serviço Suspensão', 'Valor Serviço Injeção/Ignição', 
            'Valor Serviço Cabeçote Motor Arrefecimento', 'Valor Outros Serviços', 
            'Valor Serviços Óleos', 'Valor Serviço Transmissão', 'Observações'
        ]
        sheet.append(colunas)

        # Adicionando os dados
        for ordem_servico in dados:
            linha = [
                ordem_servico['placa'],
                ordem_servico['modelo_veiculo'],
                ordem_servico['data_orcamento'],
                ordem_servico['data_faturamento'],
                ordem_servico['dias_servico'],
                ordem_servico['numero_os'],
                ordem_servico['companhia'],
                ordem_servico['valor_pecas'],
                ordem_servico['valor_servicos'],
                ordem_servico['total_os'],
                ordem_servico['valor_revitalizacao'],
                ordem_servico['valor_aditivo'],
                ordem_servico['quantidade_litros'],
                ordem_servico['valor_fluido_sangria'],
                ordem_servico['valor_palheta'],
                ordem_servico['valor_limpeza_freios'],
                ordem_servico['valor_pastilha_parabrisa'],
                ordem_servico['valor_filtro'],
                ordem_servico['valor_pneu'],
                ordem_servico['valor_bateria'],
                ordem_servico['modelo_bateria'],
                ordem_servico['lts_oleo_motor'],
                ordem_servico['valor_lt_oleo'],
                ordem_servico['marca_e_tipo_oleo'],
                ordem_servico['mecanico_servico'],
                ordem_servico['servico_filtro'],
                ordem_servico['valor_p_meta'],
                ordem_servico['valor_em_dinheiro'],
                ordem_servico['valor_servico_freios'],
                ordem_servico['valor_servico_suspensao'],
                ordem_servico['valor_servico_injecao_ignicao'],
                ordem_servico['valor_servico_cabecote_motor_arr'],
                ordem_servico['valor_outros_servicos'],
                ordem_servico['valor_servicos_oleos'],
                ordem_servico['valor_servico_transmissao'],
                ordem_servico['obs']
            ]
            sheet.append(linha)

        # Criando um buffer de memória para o Excel
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        # Criando um nome seguro para o arquivo
        nome_arquivo = f"Relatorio_{mes}_{ano}_".replace(" ", "_").replace("/", "-")

        # Retornando o Excel como resposta Flask
        response = Response(
            excel_buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.xlsx"'

        return response
    

    @app.route('/relatorio_notas', methods=['GET', 'POST'])
    def gerar_relatorio_notas():
        # Obtendo os valores do formulário
        mes = request.form.get('mes')
        ano = request.form.get('ano')
        tipo_despesa = request.form.get('tipo_despesa')
        db = dados_notas.DadosGastos()
       

        # Inicializando os objetos para acessar os dados
        if session['empresa'] == 1:
            nome_empresa = "GR7 Centro Automotivo"
        elif session['empresa'] == 2:
            nome_empresa = "Portal do Morumbi Centro Automotivo"
        elif session['empresa'] == 3:
            nome_empresa = "GR7 Morumbi Centro Automotivo"
        

        # Buscando os dados de acordo com os filtros
        dados = db.buscar_notas(mes, ano)

        # Criando o arquivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Relatório Notas Fiscais"

        # Adicionando o "header" com nome da empresa e data/hora de geração
        data_hora_geracao = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Nome da empresa (primeira linha)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        cell_empresa = sheet.cell(row=1, column=1)
        cell_empresa.value = nome_empresa
        cell_empresa.font = Font(bold=True, size=14)
        cell_empresa.alignment = Alignment(horizontal="center")

        # Data e hora de geração (segunda linha)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
        cell_data_hora = sheet.cell(row=2, column=1)
        cell_data_hora.value = f"Relatório gerado em: {data_hora_geracao}"
        cell_data_hora.font = Font(italic=True, size=10)
        cell_data_hora.alignment = Alignment(horizontal="center")

        # Adicionando cabeçalhos personalizados (começam na terceira linha)
        colunas = [
            'Pago Por', 'Emitido Para', 'Status', 'Boleto', 'Número Nota', 
            'Fornecedor', 'Data Emissão', 'Valor', 'Duplicata', 'Tipo Despesa', 'Observações'
        ]

        # Estilizando o cabeçalho
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, column_title in enumerate(colunas, 1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Ajustando largura das colunas
        for col_num, column_title in enumerate(colunas, 1):
            sheet.column_dimensions[sheet.cell(row=3, column=col_num).column_letter].width = 20

        # Adicionando os dados (a partir da linha 4)
        for row_num, nota in enumerate(dados, 4):  # Começa na linha 4
            linha = [
                nota['pago_por'],
                nota['emitido_para'],
                nota['status'],
                nota['boleto'],
                nota['numero_nota'],
                nota['fornecedor'],
                nota['data_emissao'],
                nota['valor'],
                nota['duplicata'],
                nota['tipo_despesa'],
                nota['obs']
            ]
            for col_num, cell_value in enumerate(linha, 1):
                sheet.cell(row=row_num, column=col_num).value = cell_value

        # Criando um buffer de memória para o Excel
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        # Criando um nome seguro para o arquivo
        nome_arquivo = f"Relatorio_Notas_{mes}_{ano}".replace(" ", "_").replace("/", "-")

        # Retornando o Excel como resposta Flask
        response = Response(
            excel_buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.xlsx"'

        return response


    @app.route('/relatorio_boletos', methods=['GET', 'POST'])
    def gerar_relatorio_boletos():
        # Obtendo os valores do formulário
        mes = request.form.get('mes')
        ano = request.form.get('ano')
        db = dados_notas.DadosGastos()
        # Inicializando os objetos para acessar os dados
        if session['empresa'] == 1:
            nome_empresa = "GR7 Centro Automotivo"
        elif session['empresa'] == 2:
            nome_empresa = "Portal do Morumbi Centro Automotivo"
        elif session['empresa'] == 3:
            nome_empresa = "GR7 Morumbi Centro Automotivo"

        # Buscando os dados de acordo com os filtros
        boletos = db.buscar_boletos(mes, ano)

        # Criando o arquivo Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Relatório de boletos"

        # Adicionando o "header" com nome da empresa e data/hora de geração
        
        data_hora_geracao = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Nome da empresa (primeira linha)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        cell_empresa = sheet.cell(row=1, column=1)
        cell_empresa.value = nome_empresa
        cell_empresa.font = Font(bold=True, size=14)
        cell_empresa.alignment = Alignment(horizontal="center")

        # Data e hora de geração (segunda linha)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        cell_data_hora = sheet.cell(row=2, column=1)
        cell_data_hora.value = f"Relatório gerado em: {data_hora_geracao}"
        cell_data_hora.font = Font(italic=True, size=10)
        cell_data_hora.alignment = Alignment(horizontal="center")

        # Adicionando cabeçalhos personalizados (começam na terceira linha)
        colunas = ['Número Nota', 'Notas', 'Fornecedor', 'Data Vencimento', 'Valor']

        # Estilizando o cabeçalho
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, column_title in enumerate(colunas, 1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Ajustando largura das colunas
        for col_num, column_title in enumerate(colunas, 1):
            sheet.column_dimensions[sheet.cell(row=3, column=col_num).column_letter].width = 20

        # Adicionando os dados (a partir da linha 4)
        for row_num, boleto in enumerate(boletos, 4):  # Começa na linha 4
            linha = [
                boleto['num_nota'],
                boleto['notas'],
                boleto['fornecedor'],
                boleto['data_vencimento'],
                boleto['valor']
            ]
            for col_num, cell_value in enumerate(linha, 1):
                sheet.cell(row=row_num, column=col_num).value = cell_value

        # Criando um buffer de memória para o Excel
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        # Criando um nome seguro para o arquivo
        nome_arquivo = f"Relatorio_Boletos_{mes}_{ano}".replace(" ", "_").replace("/", "-")

        # Retornando o Excel como resposta Flask
        response = Response(
            excel_buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{nome_arquivo}.xlsx"'

        return response
    
    @app.route('/gerencial')
    def tela_gerencial():
        funcionarios = [
            {'id': 1, 'nome': 'Funcionário 1'},
            {'id': 2, 'nome': 'Funcionário 2'},
            # Adicione mais funcionários conforme necessário
        ]
        return render_template('gerencial.html', empresa="Oficina", funcionarios=funcionarios)

    @app.route('/dados-loja/<loja>')
    def dados_loja(loja):
        ano = request.args.get('ano', default=2025, type=int)
        db = conection.Database() 
        faturamento = db.faturamento_loja_ano(loja, ano)
        
        

        return jsonify(faturamento)

    @app.route('/funcionarios-por-loja/<loja>')
    def funcionarios_por_loja(loja):
       
        if loja == 'GR7':
            db = conection.Database() 
        
        
        

        funcionarios = db.funcionarios_por_loja(loja)
        
        

        return jsonify(funcionarios)

    @app.route('/dados-funcionario/<mecanico>')
    def dados_funcionario(mecanico):
        loja = request.args.get('loja', default="GR7", type=str)
        ano = request.args.get('ano', default=2024, type=int)

        if loja == 'GR7':
            db = conection.Database() 
        elif loja == "Portal":
            db = conection.DatabasePortal()
        elif loja == "Morumbi":
            db = conection.DatabaseMorumbi()
        
        

        desempenho = db.desempenho_funcionario_ano(loja, mecanico, ano)
        
      

        return jsonify(desempenho)
